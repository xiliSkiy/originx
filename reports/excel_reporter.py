# -*- coding: utf-8 -*-
"""
Excel 报告生成器
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .base import BaseReporter, ReportData


class ExcelReporter(BaseReporter):
    """Excel 报告生成器"""
    
    format = "excel"
    extension = ".xlsx"
    
    # 问题类型中文名
    ISSUE_NAMES = {
        "normal": "正常",
        "blur": "图像模糊",
        "over_bright": "过度曝光",
        "under_bright": "曝光不足",
        "low_contrast": "对比度过低",
        "high_contrast": "对比度过高",
        "color_cast": "色彩偏差",
        "desaturated": "色彩饱和度低",
        "noise": "噪声干扰",
        "stripe": "条纹干扰",
        "occlusion": "画面遮挡",
        "signal_loss": "信号丢失",
        "freeze": "画面冻结",
        "scene_change": "场景变换异常",
        "shake": "视频抖动",
    }
    
    # 样式定义
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    NORMAL_FILL = PatternFill(start_color="F0F9EB", end_color="F0F9EB", fill_type="solid")
    ABNORMAL_FILL = PatternFill(start_color="FEF0F0", end_color="FEF0F0", fill_type="solid")
    
    THIN_BORDER = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    
    def generate(self, data: ReportData, output_path: str) -> str:
        """生成 Excel 报告"""
        output_file = self._ensure_output_dir(output_path)
        output_file = output_file.with_suffix(self.extension)
        
        wb = Workbook()
        
        # 摘要 Sheet
        self._create_summary_sheet(wb, data)
        
        # 异常详情 Sheet
        if data.abnormal_count > 0:
            self._create_abnormal_sheet(wb, data)
        
        # 全部结果 Sheet
        self._create_all_results_sheet(wb, data)
        
        # 删除默认 Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        wb.save(output_file)
        return str(output_file)
    
    def _create_summary_sheet(self, wb: Workbook, data: ReportData):
        """创建摘要 Sheet"""
        ws = wb.create_sheet("摘要", 0)
        
        # 标题
        ws.merge_cells("A1:C1")
        ws["A1"] = data.title
        ws["A1"].font = Font(bold=True, size=16)
        ws["A1"].alignment = Alignment(horizontal="center")
        
        # 基本信息
        info_data = [
            ("检测时间", data.timestamp.strftime("%Y-%m-%d %H:%M:%S")),
            ("检测总数", data.total_count),
            ("正常数量", data.normal_count),
            ("异常数量", data.abnormal_count),
            ("异常率", f"{data.abnormal_rate:.1f}%"),
        ]
        
        row = 3
        for label, value in info_data:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = value
            row += 1
        
        # 问题类型分布
        row += 1
        ws[f"A{row}"] = "异常类型分布"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1
        
        distribution = data.get_issue_distribution()
        for issue_type, count in distribution.items():
            ws[f"A{row}"] = self.ISSUE_NAMES.get(issue_type, issue_type)
            ws[f"B{row}"] = count
            row += 1
        
        # 调整列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
    
    def _create_abnormal_sheet(self, wb: Workbook, data: ReportData):
        """创建异常详情 Sheet"""
        ws = wb.create_sheet("异常详情")
        
        # 表头
        headers = ["文件名", "主要问题", "严重程度", "处理耗时(ms)", "建议"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # 数据
        row = 2
        for r in data.get_abnormal_results():
            filename = r.get("image_path") or r.get("video_path") or "-"
            issue = self.ISSUE_NAMES.get(r.get("primary_issue", ""), r.get("primary_issue", "-"))
            severity = r.get("severity", "-")
            process_time = r.get("total_process_time_ms") or r.get("process_time_ms") or 0
            
            # 获取建议
            suggestions = []
            for det in r.get("detection_results", []):
                if det.get("is_abnormal") and det.get("suggestions"):
                    suggestions.extend(det["suggestions"][:1])
            suggestion_text = "; ".join(suggestions[:2]) if suggestions else "-"
            
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value=issue)
            ws.cell(row=row, column=3, value=severity)
            ws.cell(row=row, column=4, value=round(process_time, 1))
            ws.cell(row=row, column=5, value=suggestion_text)
            
            # 设置样式
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.fill = self.ABNORMAL_FILL
                cell.border = self.THIN_BORDER
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 50
    
    def _create_all_results_sheet(self, wb: Workbook, data: ReportData):
        """创建全部结果 Sheet"""
        ws = wb.create_sheet("全部结果")
        
        # 表头
        headers = ["文件名", "状态", "主要问题", "处理耗时(ms)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # 数据
        row = 2
        for r in data.results:
            filename = r.get("image_path") or r.get("video_path") or "-"
            is_abnormal = r.get("is_abnormal", False)
            status = "异常" if is_abnormal else "正常"
            issue = self.ISSUE_NAMES.get(r.get("primary_issue", ""), r.get("primary_issue", "-"))
            process_time = r.get("total_process_time_ms") or r.get("process_time_ms") or 0
            
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value=status)
            ws.cell(row=row, column=3, value=issue if is_abnormal else "-")
            ws.cell(row=row, column=4, value=round(process_time, 1))
            
            # 设置样式
            fill = self.ABNORMAL_FILL if is_abnormal else self.NORMAL_FILL
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = self.THIN_BORDER
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

